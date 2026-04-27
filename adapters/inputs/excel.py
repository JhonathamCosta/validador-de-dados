from typing import Any, Dict, List

from .base import BaseInputAdapter


class ExcelInputAdapter(BaseInputAdapter):
    def __init__(self, bundle_key: str = "dados", sheet_name: str | None = None) -> None:
        self.bundle_key = bundle_key
        self.sheet_name = sheet_name

    def load(self, source: str, context: Dict[str, Any] | None = None) -> Dict[str, Any]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ImportError("openpyxl is required to load Excel files") from exc

        workbook = load_workbook(filename=source, data_only=True, read_only=True)

        if self.sheet_name:
            if self.sheet_name not in workbook.sheetnames:
                raise ValueError(f"Sheet not found: {self.sheet_name}")
            sheet = workbook[self.sheet_name]
        else:
            sheet = workbook[workbook.sheetnames[0]]

        rows_iter = sheet.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            workbook.close()
            return {self.bundle_key: []}

        headers: List[str] = []
        for idx, value in enumerate(header_row, start=1):
            if value is None:
                headers.append(f"column_{idx}")
            else:
                headers.append(str(value).strip())

        records = []
        for row in rows_iter:
            row_dict = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
            records.append(row_dict)

        workbook.close()
        return {self.bundle_key: records}
